from fastapi import FastAPI, Request, File, UploadFile, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
import os
import shutil
import json
import openpyxl
from io import BytesIO
from datetime import datetime
import logging
from typing import Optional
from .services.campaign_manager import MailCampaign

app = FastAPI(title="Mail Sender Admin Panel")

SECRET_KEY = "hfundiscabgfe76w9g32ux0jym7203"  # Debe coincidir con main.py

logger = logging.getLogger(__name__)

CLICKS_FILE = os.path.join("data", "clicks.json")

def log_click(email: str):
    try:
        data = []
        if os.path.exists(CLICKS_FILE):
            with open(CLICKS_FILE, "r") as f:
                data = json.load(f)
        
        data.append({
            "email": email,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        
        with open(CLICKS_FILE, "w") as f:
            json.dump(data, f, indent=4)
    except Exception as e:
        logger.error(f"Error logging click for {email}: {str(e)}")

# Mount the static directory
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def read_index():
    return FileResponse(os.path.join("static", "index.html"))


@app.get("/token/{token}")
async def verify_token(token: str, email: str): 
    logger.info(f"Click detectado - Email: {email}")
    
    # Registrar el click
    if email and '@' in email:
        log_click(email)
    
    # Mostrar la página de phishing (index.html)
    return FileResponse(os.path.join("static", "index.html"))

@app.post("/verify-email")
async def verify_email_endpoint(data: dict):
    email = data.get('email', '')
    return {
        "token_valid": True,
        "email": email
    }

@app.get("/health")
async def health():
    return {"status": "ok"}

# --- ADMIN ROUTES ---

ADMIN_URL = "admin-console"

@app.get(f"/{ADMIN_URL}")
async def admin_panel():
    return FileResponse(os.path.join("static", "admin.html"))

@app.post("/admin/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")
    
    file_path = os.path.join("data", "destinatarios.xlsx") # Default name used by MailCampaign
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    return {"status": "success", "filename": file.filename}

@app.post("/admin/send-campaign")
async def send_campaign(request: Request):
    try:
        # Detect base URL from request (e.g., http://localhost:8888)
        base_url = f"{request.url.scheme}://{request.url.netloc}"
        logger.info(f"Starting campaign with base URL: {base_url}")
        
        campaign = MailCampaign(base_url=base_url)
        stats = campaign.run()
        return {"status": "success", "stats": stats}
    except Exception as e:
        logger.error(f"Error running campaign: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/send-followup")
async def send_followup(data: dict):
    message_text = data.get("message", "").strip()
    if not message_text:
        raise HTTPException(status_code=400, detail="El mensaje no puede estar vacío.")

    if not os.path.exists(CLICKS_FILE):
        return {"status": "success", "message": "No hay usuarios para notificar.", "sent": 0}

    try:
        with open(CLICKS_FILE, "r") as f:
            clicks = json.load(f)
        
        # Obtener emails únicos
        unique_emails = list(set(click["email"] for click in clicks))
        
        if not unique_emails:
            return {"status": "success", "message": "No hay emails únicos.", "sent": 0}

        # Inicializar campaign para obtener credenciales SMTP
        campaign = MailCampaign()
        from .services.email_sender import EmailSender
        sender = EmailSender(
            campaign.smtp_server,
            campaign.smtp_port,
            campaign.sender_email,
            campaign.sender_password
        )

        sent_count = 0
        for email in unique_emails:
            # Crear un HTML simple para el feedback
            html_content = f"""
            <html>
                <body style="font-family: sans-serif; line-height: 1.6; color: #333;">
                    <h2>Aviso de Seguridad</h2>
                    <p>{message_text}</p>
                    <hr>
                    <p style="font-size: 0.8rem; color: #777;">Este es un mensaje automático de seguimiento de tu prueba de seguridad.</p>
                </body>
            </html>
            """
            
            success = sender.send_email(
                email,
                "Seguimiento: Prueba de Seguridad",
                html_content
            )
            if success:
                sent_count += 1
        
        return {"status": "success", "sent": sent_count, "total": len(unique_emails)}

    except Exception as e:
        logger.error(f"Error sending follow-up: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/stats")
async def get_stats():
    if os.path.exists(CLICKS_FILE):
        with open(CLICKS_FILE, "r") as f:
            return json.load(f)
    return []

@app.get("/admin/export-results")
async def export_results():
    if not os.path.exists(CLICKS_FILE):
        raise HTTPException(status_code=404, detail="No hay datos registrados para exportar.")

    try:
        with open(CLICKS_FILE, "r") as f:
            clicks = json.load(f)

        if not clicks:
            raise HTTPException(status_code=404, detail="El registro de clics está vacío.")

        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados Phishing"

        # Encabezados
        headers = ["Email", "Fecha y Hora del Clic"]
        ws.append(headers)

        # Estilo para encabezados
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Agregar datos
        for click in clicks:
            ws.append([click.get("email"), click.get("timestamp")])

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"reporte_phishing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Error exporting results: {str(e)}")
        raise HTTPException(status_code=500, detail="Error al generar el archivo Excel.")

@app.post("/admin/reset-stats")
async def reset_stats():
    try:
        if os.path.exists(CLICKS_FILE):
            os.remove(CLICKS_FILE)
            return {"status": "success", "message": "Estadísticas borradas correctamente."}
        return {"status": "success", "message": "No había datos que borrar."}
    except Exception as e:
        logger.error(f"Error resetting stats: {str(e)}")
        raise HTTPException(status_code=500, detail="Error al borrar los datos.")

@app.get("/admin/download-template")
async def download_template():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Destinatarios"

        # Encabezados
        ws.append(["Nombre", "Email"])
        ws.append(["Juan Perez", "juan.perez@ejemplo.com"])
        ws.append(["Maria Garcia", "m.garcia@empresa.com"])

        # Ajustar ancho
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "plantilla_campaña_seguridad.xlsx"
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error generating template: {str(e)}")
        raise HTTPException(status_code=500, detail="Error al generar la plantilla.")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8888)