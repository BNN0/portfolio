import openpyxl
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)

class ExcelReader:
    
    @staticmethod
    def read_emails(file_path: str, column_name: str = "email") -> List[str]:
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Obtener los encabezados
            headers = [cell.value for cell in worksheet[1]]
            
            if column_name not in headers:
                raise ValueError(f"Columna '{column_name}' no encontrada. Disponibles: {headers}")
            
            # Obtener índice de la columna
            column_index = headers.index(column_name) + 1
            
            # Leer emails
            emails = []
            for row in worksheet.iter_rows(min_row=2, values_only=False):
                cell_value = row[column_index - 1].value
                if cell_value:
                    email_str = str(cell_value).strip()
                    # Validar formato básico de email
                    if "@" in email_str and "." in email_str:
                        emails.append(email_str)
                    else:
                        logger.warning(f"Email inválido ignorado: {email_str}")
            
            # Remover duplicados manteniendo orden
            unique_emails = []
            seen = set()
            for email in emails:
                if email not in seen:
                    unique_emails.append(email)
                    seen.add(email)
            
            logger.info(f"Se leyeron {len(unique_emails)} emails únicos de {file_path}")
            return unique_emails
            
        except FileNotFoundError:
            logger.error(f"Archivo no encontrado: {file_path}")
            raise
        except Exception as e:
            logger.error(f"Error leyendo Excel: {str(e)}")
            raise
    
    @staticmethod
    def read_emails_with_data(file_path: str, email_column: str = "email") -> List[Dict[str, any]]:
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Obtener encabezados
            headers = [cell.value for cell in worksheet[1]]
            
            if email_column not in headers:
                raise ValueError(f"Columna '{email_column}' no encontrada")
            
            # Leer datos
            rows_data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if row_dict.get(email_column):
                    rows_data.append(row_dict)
            
            logger.info(f"Se leyeron {len(rows_data)} registros de {file_path}")
            return rows_data
            
        except Exception as e:
            logger.error(f"Error leyendo Excel con datos: {str(e)}")
            raise